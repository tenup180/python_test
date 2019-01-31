import glob
import openpyxl


wb=openpyxl.load_workbook("example_1.xlsx")
sheet=wb['Sheet1']
values= sheet.cell(row=1, column=1).value
print(values)

for cell_obj in list(sheet.columns)[1]:
    print(cell_obj.value)
"""   
new_wb=openpyxl.Workbook()
new_wb.create_sheet(index=3,title="test30sheet")
new_wb.save("test.xlsx")"""
wb.remove_sheet(wb.get_sheet_by_name("test30sheet"))
a=wb.get_sheet_names()
print(a)
